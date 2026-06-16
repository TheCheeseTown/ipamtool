import csv
import ipaddress
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .storage import Subnet
from .ip_utils import subnet_tree

HEADERS = [
    "VLAN ID", "Segment Name", "Device Name", "Purpose", "Subnet", "CIDR",
    "Total Addresses", "Usable Addresses",
    "Gateway", "DHCP Start", "DHCP End", "Static Range",
    "Location", "Owner", "Status", "Notes", "Parent Range", "Parent Subnet",
]


def _address_counts(s: Subnet) -> tuple[int, int]:
    net = s.network
    total = net.num_addresses
    usable = len(list(net.hosts()))
    return total, usable

STATUS_COLORS = {
    "active":     "C6EFCE",  # green
    "reserved":   "FFEB9C",  # yellow
    "planned":    "BDD7EE",  # blue
    "deprecated": "FFC7CE",  # red
}

# Indent per depth level in Excel (em-spaces)
_INDENT = "    "


def _gw_lookup(subnets: list[Subnet]) -> dict[str, str]:
    return {s.cidr_notation(): s.gateway for s in subnets if not s.is_host and s.gateway}


def _effective_gw(s: Subnet, gw_by_cidr: dict[str, str]) -> str:
    if s.gateway:
        return s.gateway
    if s.is_host and s.parent_subnet:
        return gw_by_cidr.get(s.parent_subnet, "")
    return ""


def _row(s: Subnet, depth: int = 0, gw_by_cidr: dict | None = None) -> list:
    indent = _INDENT * depth
    total, usable = _address_counts(s)
    gw = _effective_gw(s, gw_by_cidr or {})
    return [
        s.vlan_id,
        indent + (s.segment_name or ""),
        s.device_name,
        s.purpose,
        s.subnet,
        s.cidr,
        total,
        usable,
        gw,
        s.dhcp_start,
        s.dhcp_end,
        s.static_range,
        s.location,
        s.owner,
        s.status,
        s.notes,
        s.parent_range,
        s.parent_subnet,
    ]


def export_csv(subnets: list[Subnet], path: Path):
    gw_by_cidr = _gw_lookup(subnets)
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for s, depth in subnet_tree(subnets):
            writer.writerow(_row(s, depth, gw_by_cidr))


def export_xlsx(subnets: list[Subnet], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "IPAM"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_alignment = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    gw_by_cidr = _gw_lookup(subnets)
    for s, depth in subnet_tree(subnets):
        ws.append(_row(s, depth, gw_by_cidr))
        row_num = ws.max_row

        # Indent the Segment Name cell using openpyxl indent level
        name_cell = ws.cell(row=row_num, column=2)
        name_cell.alignment = Alignment(indent=depth * 2)

        # Dim child rows slightly
        if depth > 0:
            dim_font = Font(color="555555" if depth > 1 else "222222")
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_num, column=col).font = dim_font

        # Status color fill
        status = s.status.lower()
        color = STATUS_COLORS.get(status)
        if color:
            fill = PatternFill(fill_type="solid", fgColor=color)
            for cell in ws[row_num]:
                cell.fill = fill

        # Bold top-level rows
        if depth == 0:
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = Font(bold=True)

    for col_idx, _ in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(path)
